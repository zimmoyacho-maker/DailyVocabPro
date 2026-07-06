#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Import DailyVocab Master Dictionary Excel into SQLite."""

from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path
from typing import Any, List

from openpyxl import load_workbook

SCHEMA_PATH = Path(__file__).resolve().parent / "schema.sql"


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").replace("\u2003", " ").strip()


def split_tags(tags: str) -> List[str]:
    return [x.strip() for x in tags.replace(";", ",").split(",") if x.strip()]


def connect(db_path: Path) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(SCHEMA_PATH.read_text(encoding="utf-8"))
    conn.commit()


def import_excel(input_path: Path, db_path: Path) -> None:
    wb = load_workbook(input_path)
    ws = wb["words"] if "words" in wb.sheetnames else wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        name = norm(ws.cell(1, col).value)
        if name:
            headers[name] = col

    missing = [c for c in ["word", "meaning", "example"] if c not in headers]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    conn = connect(db_path)
    init_db(conn)
    cur = conn.cursor()

    imported = 0
    for row in range(2, ws.max_row + 1):
        word = norm(ws.cell(row, headers["word"]).value)
        if not word:
            continue

        def get(col_name: str) -> str:
            col = headers.get(col_name)
            return norm(ws.cell(row, col).value) if col else ""

        meaning = get("meaning")
        example_en = get("example")
        example_ko = get("example_ko")
        memo = get("memo")
        level = get("level")
        tags = get("tags")
        part_of_speech = get("part_of_speech")
        pronunciation = get("pronunciation")

        cur.execute("""
            INSERT INTO words(word, meaning, part_of_speech, pronunciation, level, source, updated_at)
            VALUES(?,?,?,?,?,?,CURRENT_TIMESTAMP)
            ON CONFLICT(word) DO UPDATE SET
                meaning=excluded.meaning,
                part_of_speech=excluded.part_of_speech,
                pronunciation=excluded.pronunciation,
                level=excluded.level,
                source=excluded.source,
                updated_at=CURRENT_TIMESTAMP
        """, (word, meaning, part_of_speech, pronunciation, level, input_path.name))

        word_id = cur.execute("SELECT id FROM words WHERE word=?", (word,)).fetchone()[0]

        existing = cur.execute(
            "SELECT id FROM examples WHERE word_id=? AND is_primary=1", (word_id,)
        ).fetchone()
        if existing:
            cur.execute("""
                UPDATE examples
                SET example_en=?, example_ko=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            """, (example_en, example_ko, existing[0]))
        else:
            cur.execute("""
                INSERT INTO examples(word_id, example_en, example_ko, is_primary)
                VALUES(?,?,?,1)
            """, (word_id, example_en, example_ko))

        cur.execute("""
            INSERT INTO notes(word_id, memo, updated_at)
            VALUES(?,?,CURRENT_TIMESTAMP)
            ON CONFLICT(word_id) DO UPDATE SET memo=excluded.memo, updated_at=CURRENT_TIMESTAMP
        """, (word_id, memo))

        cur.execute("DELETE FROM word_tags WHERE word_id=?", (word_id,))
        for tag in split_tags(tags):
            cur.execute("INSERT OR IGNORE INTO word_tags(word_id, tag) VALUES(?,?)", (word_id, tag))

        imported += 1

    conn.commit()
    conn.close()
    print(f"Imported {imported} words into {db_path}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--db", required=True)
    args = parser.parse_args()
    import_excel(Path(args.input), Path(args.db))


if __name__ == "__main__":
    main()
