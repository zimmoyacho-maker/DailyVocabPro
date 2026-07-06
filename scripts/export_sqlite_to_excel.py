#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Export DailyVocab SQLite DB back to Excel."""

from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

COLUMNS = [
    "word", "meaning", "example", "example_ko", "memo", "level", "tags",
    "part_of_speech", "pronunciation", "grammar", "collocations", "synonyms",
    "antonyms", "confusion", "memory_tip", "ai_reviewed", "human_reviewed",
]


def export_excel(db_path: Path, output_path: Path) -> None:
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("""
        SELECT w.id, w.word, w.meaning,
               COALESCE(e.example_en,''), COALESCE(e.example_ko,''),
               COALESCE(n.memo,''), COALESCE(w.level,''),
               COALESCE(w.part_of_speech,''), COALESCE(w.pronunciation,''),
               COALESCE(n.grammar,''), COALESCE(n.collocations,''),
               COALESCE(n.synonyms,''), COALESCE(n.antonyms,''),
               COALESCE(n.confusion,''), COALESCE(n.memory_tip,''),
               COALESCE(n.ai_reviewed,0), COALESCE(n.human_reviewed,0)
        FROM words w
        LEFT JOIN examples e ON e.word_id=w.id AND e.is_primary=1
        LEFT JOIN notes n ON n.word_id=w.id
        ORDER BY w.id
    """)
    rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "words"
    ws.append(COLUMNS)

    for item in rows:
        word_id = item[0]
        tags = ", ".join(r[0] for r in cur.execute(
            "SELECT tag FROM word_tags WHERE word_id=? ORDER BY tag", (word_id,)
        ).fetchall())
        ws.append([
            item[1], item[2], item[3], item[4], item[5], item[6], tags,
            item[7], item[8], item[9], item[10], item[11], item[12],
            item[13], item[14], item[15], item[16],
        ])

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {"A":18,"B":24,"C":54,"D":54,"E":56,"F":10,"G":22,"H":16,"I":18,
              "J":36,"K":36,"L":28,"M":28,"N":28,"O":36,"P":12,"Q":14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    conn.close()
    print(f"Exported {len(rows)} rows to {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    export_excel(Path(args.db), Path(args.output))


if __name__ == "__main__":
    main()
