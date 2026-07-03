#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DailyVocabPro Master Dictionary Editor v1.0

Run:
    python apps/dictionary_editor.py

Required:
    pip install openpyxl
"""

from __future__ import annotations

import shutil
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


REQUIRED_COLUMNS = ["word", "meaning", "example"]
EDIT_COLUMNS = ["example_ko", "memo", "level", "tags"]
OPTIONAL_COLUMNS = ["qa_flag"]
ALL_COLUMNS = REQUIRED_COLUMNS + EDIT_COLUMNS + OPTIONAL_COLUMNS

LEVEL_OPTIONS = ["", "A1", "A2", "B1", "B2", "C1", "C2"]
TAG_OPTIONS = [
    "Daily",
    "Business",
    "Academic",
    "Education",
    "Travel",
    "Medical",
    "Legal",
    "Science",
    "Technology",
    "Conversation",
]


@dataclass
class RowData:
    row_index: int
    word: str
    meaning: str
    example: str
    example_ko: str
    memo: str
    level: str
    tags: str
    qa_flag: str


def norm(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").replace("\u2003", " ").strip()


class DictionaryWorkbook:
    def __init__(self) -> None:
        self.path: Optional[Path] = None
        self.wb = None
        self.ws = None
        self.headers: Dict[str, int] = {}
        self.data_rows: List[int] = []
        self.current_pos: int = 0

    def load(self, path: Path) -> None:
        self.path = path
        self.wb = load_workbook(path)
        self.ws = self.wb["words"] if "words" in self.wb.sheetnames else self.wb.active
        self.headers = self._ensure_columns()
        self.data_rows = [
            row for row in range(2, self.ws.max_row + 1)
            if norm(self.ws.cell(row=row, column=self.headers["word"]).value)
        ]
        self.current_pos = 0
        self._style_sheet()

    def _header_map(self) -> Dict[str, int]:
        headers = {}
        for col in range(1, self.ws.max_column + 1):
            name = norm(self.ws.cell(row=1, column=col).value)
            if name:
                headers[name] = col
        return headers

    def _ensure_columns(self) -> Dict[str, int]:
        headers = self._header_map()
        missing = [c for c in REQUIRED_COLUMNS if c not in headers]
        if missing:
            raise ValueError(f"필수 컬럼이 없습니다: {missing}")

        next_col = self.ws.max_column + 1
        for col_name in ALL_COLUMNS:
            if col_name not in headers:
                self.ws.cell(row=1, column=next_col).value = col_name
                headers[col_name] = next_col
                next_col += 1
        return headers

    def _style_sheet(self) -> None:
        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(bold=True, color="FFFFFF")
        for cell in self.ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        widths = {
            "word": 18,
            "meaning": 24,
            "example": 54,
            "example_ko": 54,
            "memo": 56,
            "level": 10,
            "tags": 22,
            "qa_flag": 28,
        }
        for name, col in self.headers.items():
            self.ws.column_dimensions[get_column_letter(col)].width = widths.get(name, 18)
        self.ws.freeze_panes = "A2"

    def count(self) -> int:
        return len(self.data_rows)

    def get_current(self) -> RowData:
        if not self.data_rows:
            raise RuntimeError("데이터가 없습니다.")
        row = self.data_rows[self.current_pos]
        return self.get_row(row)

    def get_row(self, row: int) -> RowData:
        h = self.headers
        return RowData(
            row_index=row,
            word=norm(self.ws.cell(row=row, column=h["word"]).value),
            meaning=norm(self.ws.cell(row=row, column=h["meaning"]).value),
            example=norm(self.ws.cell(row=row, column=h["example"]).value),
            example_ko=norm(self.ws.cell(row=row, column=h["example_ko"]).value),
            memo=norm(self.ws.cell(row=row, column=h["memo"]).value),
            level=norm(self.ws.cell(row=row, column=h["level"]).value),
            tags=norm(self.ws.cell(row=row, column=h["tags"]).value),
            qa_flag=norm(self.ws.cell(row=row, column=h["qa_flag"]).value),
        )

    def update_current(self, example_ko: str, memo: str, level: str, tags: str) -> None:
        row = self.data_rows[self.current_pos]
        h = self.headers
        self.ws.cell(row=row, column=h["example_ko"]).value = example_ko.strip()
        self.ws.cell(row=row, column=h["memo"]).value = memo.strip()
        self.ws.cell(row=row, column=h["level"]).value = level.strip()
        self.ws.cell(row=row, column=h["tags"]).value = tags.strip()
        self.ws.cell(row=row, column=h["qa_flag"]).value = self._qa_flag(row)

    def _qa_flag(self, row: int) -> str:
        h = self.headers
        flags = []
        if not norm(self.ws.cell(row=row, column=h["example_ko"]).value):
            flags.append("missing_example_ko")
        if not norm(self.ws.cell(row=row, column=h["memo"]).value):
            flags.append("missing_memo")
        if not norm(self.ws.cell(row=row, column=h["level"]).value):
            flags.append("missing_level")
        if not norm(self.ws.cell(row=row, column=h["tags"]).value):
            flags.append("missing_tags")
        example_ko = norm(self.ws.cell(row=row, column=h["example_ko"]).value)
        if example_ko and len(example_ko) < 8:
            flags.append("short_translation")
        return ", ".join(flags)

    def refresh_all_qa(self) -> Dict[str, int]:
        summary = {
            "total": 0,
            "complete": 0,
            "missing_example_ko": 0,
            "missing_memo": 0,
            "missing_level": 0,
            "missing_tags": 0,
            "short_translation": 0,
        }
        for row in self.data_rows:
            summary["total"] += 1
            flag = self._qa_flag(row)
            self.ws.cell(row=row, column=self.headers["qa_flag"]).value = flag
            if not flag:
                summary["complete"] += 1
            for key in summary:
                if key not in ("total", "complete") and key in flag:
                    summary[key] += 1
        return summary

    def search(self, query: str) -> List[int]:
        q = query.lower().strip()
        if not q:
            return []
        result = []
        h = self.headers
        fields = ["word", "meaning", "example", "example_ko", "memo", "level", "tags"]
        for pos, row in enumerate(self.data_rows):
            text = " ".join(norm(self.ws.cell(row=row, column=h[f]).value) for f in fields).lower()
            if q in text:
                result.append(pos)
        return result

    def goto_pos(self, pos: int) -> None:
        if not self.data_rows:
            return
        self.current_pos = max(0, min(pos, len(self.data_rows) - 1))

    def goto_row_number(self, row_number: int) -> None:
        if row_number in self.data_rows:
            self.current_pos = self.data_rows.index(row_number)
            return
        candidates = [r for r in self.data_rows if r >= row_number]
        if candidates:
            self.current_pos = self.data_rows.index(candidates[0])

    def append_log(self, note: str) -> None:
        if "Translation_Log" in self.wb.sheetnames:
            ws = self.wb["Translation_Log"]
        else:
            ws = self.wb.create_sheet("Translation_Log")
            ws.append(["timestamp", "tool", "note"])

        ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "dictionary_editor", note])
        fill = PatternFill("solid", fgColor="0F766E")
        font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 80

    def backup(self) -> Path:
        if not self.path:
            raise RuntimeError("파일 경로가 없습니다.")
        backup_dir = self.path.parent / "backups"
        backup_dir.mkdir(exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"{self.path.stem}_backup_{stamp}{self.path.suffix}"
        shutil.copy2(self.path, backup_path)
        return backup_path

    def save(self) -> None:
        if not self.path:
            raise RuntimeError("저장할 파일이 없습니다.")
        self._style_sheet()
        self.append_log("Saved from Master Dictionary Editor")
        self.wb.save(self.path)

    def save_as(self, path: Path) -> None:
        self.path = path
        self.save()


class DictionaryEditor(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("DailyVocab Master Dictionary Editor")
        self.geometry("1180x760")
        self.minsize(1000, 680)

        self.model = DictionaryWorkbook()
        self.search_positions: List[int] = []
        self.search_index = 0
        self.dirty = False

        self._build_ui()
        self._bind_keys()

    def _build_ui(self):
        self.columnconfigure(0, weight=3)
        self.columnconfigure(1, weight=2)
        self.rowconfigure(1, weight=1)

        toolbar = ttk.Frame(self, padding=8)
        toolbar.grid(row=0, column=0, columnspan=2, sticky="ew")

        ttk.Button(toolbar, text="Excel 열기", command=self.open_file).pack(side="left")
        ttk.Button(toolbar, text="저장", command=self.save_file).pack(side="left", padx=4)
        ttk.Button(toolbar, text="다른 이름으로 저장", command=self.save_as).pack(side="left", padx=4)
        ttk.Button(toolbar, text="QA 검사", command=self.run_qa).pack(side="left", padx=12)

        ttk.Label(toolbar, text="행 이동").pack(side="left", padx=(18, 4))
        self.goto_var = tk.StringVar()
        ttk.Entry(toolbar, width=8, textvariable=self.goto_var).pack(side="left")
        ttk.Button(toolbar, text="Go", command=self.goto_row).pack(side="left", padx=4)

        ttk.Label(toolbar, text="검색").pack(side="left", padx=(20, 4))
        self.search_var = tk.StringVar()
        ttk.Entry(toolbar, width=30, textvariable=self.search_var).pack(side="left")
        ttk.Button(toolbar, text="찾기", command=self.search).pack(side="left", padx=4)
        ttk.Button(toolbar, text="다음", command=self.search_next).pack(side="left", padx=4)

        self.status_var = tk.StringVar(value="Excel 파일을 열어 주세요.")
        ttk.Label(toolbar, textvariable=self.status_var).pack(side="right")

        left = ttk.Frame(self, padding=10)
        left.grid(row=1, column=0, sticky="nsew")
        left.columnconfigure(0, weight=1)
        left.rowconfigure(8, weight=1)

        self.word_var = tk.StringVar()
        self.meaning_var = tk.StringVar()
        self.row_var = tk.StringVar()

        ttk.Label(left, textvariable=self.row_var).grid(row=0, column=0, sticky="w")
        ttk.Label(left, textvariable=self.word_var, font=("Segoe UI", 24, "bold")).grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Label(left, textvariable=self.meaning_var, font=("Segoe UI", 12)).grid(row=2, column=0, sticky="w", pady=(0, 12))

        ttk.Label(left, text="Example", font=("Segoe UI", 11, "bold")).grid(row=3, column=0, sticky="w")
        self.example_text = self._text(left, height=6, readonly=True)
        self.example_text.grid(row=4, column=0, sticky="nsew", pady=(4, 12))

        ttk.Label(left, text="Example_KO", font=("Segoe UI", 11, "bold")).grid(row=5, column=0, sticky="w")
        self.example_ko_text = self._text(left, height=7)
        self.example_ko_text.grid(row=6, column=0, sticky="nsew", pady=(4, 12))

        ttk.Label(left, text="Memo", font=("Segoe UI", 11, "bold")).grid(row=7, column=0, sticky="w")
        self.memo_text = self._text(left, height=8)
        self.memo_text.grid(row=8, column=0, sticky="nsew", pady=(4, 12))

        nav = ttk.Frame(left)
        nav.grid(row=9, column=0, sticky="ew", pady=(4, 0))
        ttk.Button(nav, text="← 이전", command=self.prev_row).pack(side="left")
        ttk.Button(nav, text="저장 후 다음 →", command=self.save_and_next).pack(side="left", padx=8)
        ttk.Button(nav, text="다음 →", command=self.next_row).pack(side="left")

        right = ttk.Frame(self, padding=10)
        right.grid(row=1, column=1, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(7, weight=1)

        ttk.Label(right, text="Level", font=("Segoe UI", 11, "bold")).grid(row=0, column=0, sticky="w")
        self.level_var = tk.StringVar()
        self.level_combo = ttk.Combobox(right, textvariable=self.level_var, values=LEVEL_OPTIONS, state="readonly")
        self.level_combo.grid(row=1, column=0, sticky="ew", pady=(4, 12))

        ttk.Label(right, text="Tags", font=("Segoe UI", 11, "bold")).grid(row=2, column=0, sticky="w")
        self.tags_var = tk.StringVar()
        self.tags_entry = ttk.Entry(right, textvariable=self.tags_var)
        self.tags_entry.grid(row=3, column=0, sticky="ew", pady=(4, 6))

        tags_box = ttk.LabelFrame(right, text="Quick Tags", padding=8)
        tags_box.grid(row=4, column=0, sticky="ew", pady=(0, 12))
        for i, tag in enumerate(TAG_OPTIONS):
            ttk.Button(tags_box, text=tag, command=lambda t=tag: self.add_tag(t)).grid(
                row=i // 2, column=i % 2, sticky="ew", padx=3, pady=3
            )
        tags_box.columnconfigure(0, weight=1)
        tags_box.columnconfigure(1, weight=1)

        ttk.Label(right, text="QA Flag", font=("Segoe UI", 11, "bold")).grid(row=5, column=0, sticky="w")
        self.qa_var = tk.StringVar()
        ttk.Label(right, textvariable=self.qa_var, wraplength=360).grid(row=6, column=0, sticky="ew", pady=(4, 12))

        self.qa_summary = tk.Text(right, height=12, wrap="word")
        self.qa_summary.grid(row=7, column=0, sticky="nsew", pady=(8, 0))

        self._watch_dirty(self.example_ko_text)
        self._watch_dirty(self.memo_text)
        self.level_var.trace_add("write", lambda *_: self._mark_dirty())
        self.tags_var.trace_add("write", lambda *_: self._mark_dirty())

    def _text(self, parent, height=5, readonly=False):
        t = tk.Text(parent, height=height, wrap="word", font=("Segoe UI", 10), padx=8, pady=8)
        if readonly:
            t.configure(state="disabled", background="#F5F5F5")
        return t

    def _watch_dirty(self, text_widget):
        text_widget.bind("<<Modified>>", self._on_text_modified)

    def _on_text_modified(self, event):
        widget = event.widget
        if widget.edit_modified():
            self._mark_dirty()
            widget.edit_modified(False)

    def _mark_dirty(self):
        self.dirty = True

    def _bind_keys(self):
        self.bind("<Control-o>", lambda e: self.open_file())
        self.bind("<Control-s>", lambda e: self.save_file())
        self.bind("<Alt-Left>", lambda e: self.prev_row())
        self.bind("<Alt-Right>", lambda e: self.next_row())
        self.bind("<F5>", lambda e: self.run_qa())

    def open_file(self):
        path = filedialog.askopenfilename(
            title="Master Dictionary Excel 파일 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            self.model.load(Path(path))
            self.dirty = False
            self.load_current_to_ui()
            self.status_var.set(f"Loaded: {Path(path).name}")
        except Exception as e:
            messagebox.showerror("파일 열기 오류", str(e))

    def _set_text(self, widget, value: str, readonly=False):
        if readonly:
            widget.configure(state="normal")
        widget.delete("1.0", "end")
        widget.insert("1.0", value)
        if readonly:
            widget.configure(state="disabled")
        widget.edit_modified(False)

    def _get_text(self, widget) -> str:
        return widget.get("1.0", "end").strip()

    def load_current_to_ui(self):
        if not self.model.data_rows:
            return
        row = self.model.get_current()
        self.row_var.set(f"Excel Row {row.row_index}   ({self.model.current_pos + 1} / {self.model.count()})")
        self.word_var.set(row.word)
        self.meaning_var.set(row.meaning)
        self._set_text(self.example_text, row.example, readonly=True)
        self._set_text(self.example_ko_text, row.example_ko)
        self._set_text(self.memo_text, row.memo)
        self.level_var.set(row.level)
        self.tags_var.set(row.tags)
        self.qa_var.set(row.qa_flag or "OK")
        self.goto_var.set(str(row.row_index))
        self.dirty = False

    def save_current_from_ui(self):
        if not self.model.data_rows:
            return
        self.model.update_current(
            example_ko=self._get_text(self.example_ko_text),
            memo=self._get_text(self.memo_text),
            level=self.level_var.get(),
            tags=self.tags_var.get(),
        )
        row = self.model.get_current()
        self.qa_var.set(row.qa_flag or "OK")
        self.dirty = False

    def _confirm_discard_or_save(self) -> bool:
        if not self.dirty:
            return True
        result = messagebox.askyesnocancel("저장 확인", "현재 행에 저장되지 않은 변경이 있습니다. 저장할까요?")
        if result is None:
            return False
        if result:
            self.save_current_from_ui()
        return True

    def prev_row(self):
        if not self._confirm_discard_or_save():
            return
        self.model.goto_pos(self.model.current_pos - 1)
        self.load_current_to_ui()

    def next_row(self):
        if not self._confirm_discard_or_save():
            return
        self.model.goto_pos(self.model.current_pos + 1)
        self.load_current_to_ui()

    def save_and_next(self):
        self.save_current_from_ui()
        self.model.goto_pos(self.model.current_pos + 1)
        self.load_current_to_ui()

    def goto_row(self):
        if not self._confirm_discard_or_save():
            return
        try:
            row_number = int(self.goto_var.get())
        except ValueError:
            messagebox.showwarning("행 이동", "행 번호를 숫자로 입력하세요.")
            return
        self.model.goto_row_number(row_number)
        self.load_current_to_ui()

    def search(self):
        query = self.search_var.get().strip()
        if not query:
            return
        self.search_positions = self.model.search(query)
        self.search_index = 0
        if not self.search_positions:
            messagebox.showinfo("검색", "검색 결과가 없습니다.")
            return
        if not self._confirm_discard_or_save():
            return
        self.model.goto_pos(self.search_positions[0])
        self.load_current_to_ui()
        self.status_var.set(f"Search: {len(self.search_positions)} result(s)")

    def search_next(self):
        if not self.search_positions:
            self.search()
            return
        if not self._confirm_discard_or_save():
            return
        self.search_index = (self.search_index + 1) % len(self.search_positions)
        self.model.goto_pos(self.search_positions[self.search_index])
        self.load_current_to_ui()
        self.status_var.set(f"Search: {self.search_index + 1}/{len(self.search_positions)}")

    def add_tag(self, tag: str):
        existing = [x.strip() for x in self.tags_var.get().split(",") if x.strip()]
        if tag not in existing:
            existing.append(tag)
        self.tags_var.set(", ".join(existing))

    def run_qa(self):
        if not self.model.data_rows:
            return
        self.save_current_from_ui()
        summary = self.model.refresh_all_qa()
        complete_rate = int((summary["complete"] / summary["total"]) * 100) if summary["total"] else 0
        text = [
            "QA Summary",
            "----------",
            f"Total: {summary['total']}",
            f"Complete: {summary['complete']}",
            f"Missing example_ko: {summary['missing_example_ko']}",
            f"Missing memo: {summary['missing_memo']}",
            f"Missing level: {summary['missing_level']}",
            f"Missing tags: {summary['missing_tags']}",
            f"Short translation: {summary['short_translation']}",
            f"Complete rate: {complete_rate}%",
        ]
        self.qa_summary.delete("1.0", "end")
        self.qa_summary.insert("1.0", "\n".join(text))
        row = self.model.get_current()
        self.qa_var.set(row.qa_flag or "OK")

    def save_file(self):
        if not self.model.path:
            messagebox.showwarning("저장", "먼저 Excel 파일을 열어 주세요.")
            return
        try:
            self.save_current_from_ui()
            backup_path = self.model.backup()
            self.model.save()
            self.status_var.set(f"Saved. Backup: {backup_path.name}")
            messagebox.showinfo("저장 완료", f"저장되었습니다.\n백업: {backup_path.name}")
        except Exception as e:
            messagebox.showerror("저장 오류", str(e))

    def save_as(self):
        if not self.model.data_rows:
            return
        path = filedialog.asksaveasfilename(
            title="다른 이름으로 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not path:
            return
        try:
            self.save_current_from_ui()
            self.model.save_as(Path(path))
            self.status_var.set(f"Saved as: {Path(path).name}")
        except Exception as e:
            messagebox.showerror("저장 오류", str(e))


def main():
    app = DictionaryEditor()
    app.mainloop()


if __name__ == "__main__":
    main()
