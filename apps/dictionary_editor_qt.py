#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DailyVocabPro Master Dictionary Editor v2.5
Knowledge Engine Integrated
"""

from __future__ import annotations

import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from PySide6.QtCore import Qt
from PySide6.QtGui import QAction, QKeySequence
from PySide6.QtWidgets import (
    QApplication, QCheckBox, QComboBox, QFileDialog, QFrame, QGridLayout,
    QHBoxLayout, QLabel, QLineEdit, QListWidget, QListWidgetItem, QMainWindow,
    QMessageBox, QPlainTextEdit, QProgressBar, QPushButton, QSplitter,
    QStatusBar, QVBoxLayout, QWidget,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from engine.knowledge_engine import KnowledgeEngine

REQUIRED_COLUMNS = ["word", "meaning", "example"]
ALL_COLUMNS = ["word", "meaning", "example", "example_ko", "memo", "level", "tags", "qa_flag"]
LEVEL_OPTIONS = ["", "A1", "A2", "B1", "B2", "C1", "C2"]
TAG_OPTIONS = ["Daily", "Business", "Academic", "Education", "Travel", "Medical", "Legal", "Science", "Technology", "Conversation"]

DARK_STYLE = """
QMainWindow, QWidget { background-color: #111827; color: #E5E7EB; font-family: Segoe UI; font-size: 10.5pt; }
QFrame#Card { background-color: #1F2937; border: 1px solid #374151; border-radius: 10px; }
QLabel#TitleWord { font-size: 30px; font-weight: 700; color: #F9FAFB; }
QLabel#Meaning { font-size: 16px; color: #D1D5DB; }
QLabel#SectionTitle { font-size: 12px; font-weight: 700; color: #93C5FD; }
QPlainTextEdit, QLineEdit, QComboBox, QListWidget { background-color: #0B1220; color: #E5E7EB; border: 1px solid #374151; border-radius: 6px; padding: 6px; }
QPushButton { background-color: #2563EB; color: white; border: 0; border-radius: 7px; padding: 8px 12px; font-weight: 600; }
QPushButton:hover { background-color: #1D4ED8; }
QPushButton#Secondary { background-color: #374151; }
QPushButton#Assist { background-color: #7C3AED; }
QPushButton#Assist:hover { background-color: #6D28D9; }
QProgressBar { border: 1px solid #374151; border-radius: 6px; text-align: center; background-color: #0B1220; }
QProgressBar::chunk { background-color: #2563EB; border-radius: 6px; }
QStatusBar { background-color: #0B1220; color: #D1D5DB; }
"""

def norm(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").replace("\u2003", " ").strip()

class DictionaryWorkbook:
    def __init__(self) -> None:
        self.path: Optional[Path] = None
        self.wb = None
        self.ws = None
        self.headers: Dict[str, int] = {}
        self.data_rows: List[int] = []
        self.current_pos = 0

    def load(self, path: Path) -> None:
        self.path = path
        self.wb = load_workbook(path)
        self.ws = self.wb["words"] if "words" in self.wb.sheetnames else self.wb.active
        self.headers = self.ensure_columns()
        self.data_rows = [r for r in range(2, self.ws.max_row + 1) if norm(self.ws.cell(r, self.headers["word"]).value)]
        self.current_pos = 0
        self.style()

    def ensure_columns(self) -> Dict[str, int]:
        headers = {}
        for col in range(1, self.ws.max_column + 1):
            name = norm(self.ws.cell(1, col).value)
            if name:
                headers[name] = col
        missing = [c for c in REQUIRED_COLUMNS if c not in headers]
        if missing:
            raise ValueError(f"필수 컬럼이 없습니다: {missing}")
        next_col = self.ws.max_column + 1
        for c in ALL_COLUMNS:
            if c not in headers:
                self.ws.cell(1, next_col).value = c
                headers[c] = next_col
                next_col += 1
        return headers

    def style(self) -> None:
        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(bold=True, color="FFFFFF")
        for cell in self.ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths = {"word":18, "meaning":24, "example":54, "example_ko":54, "memo":56, "level":10, "tags":22, "qa_flag":28}
        for name, col in self.headers.items():
            self.ws.column_dimensions[get_column_letter(col)].width = widths.get(name, 18)
        self.ws.freeze_panes = "A2"

    def count(self): return len(self.data_rows)
    def pos_for_row(self, row): return self.data_rows.index(row) if row in self.data_rows else None
    def goto_pos(self, pos): self.current_pos = max(0, min(pos, len(self.data_rows)-1))
    def goto_row_number(self, row): 
        if row in self.data_rows: self.current_pos = self.data_rows.index(row)

    def get_current(self):
        row = self.data_rows[self.current_pos]
        h = self.headers
        return {
            "row": row,
            "word": norm(self.ws.cell(row, h["word"]).value),
            "meaning": norm(self.ws.cell(row, h["meaning"]).value),
            "example": norm(self.ws.cell(row, h["example"]).value),
            "example_ko": norm(self.ws.cell(row, h["example_ko"]).value),
            "memo": norm(self.ws.cell(row, h["memo"]).value),
            "level": norm(self.ws.cell(row, h["level"]).value),
            "tags": norm(self.ws.cell(row, h["tags"]).value),
            "qa_flag": norm(self.ws.cell(row, h["qa_flag"]).value),
        }

    def qa_flag_for_row(self, row: int) -> str:
        h = self.headers
        flags = []
        if not norm(self.ws.cell(row, h["example_ko"]).value): flags.append("missing_example_ko")
        if not norm(self.ws.cell(row, h["memo"]).value): flags.append("missing_memo")
        if not norm(self.ws.cell(row, h["level"]).value): flags.append("missing_level")
        if not norm(self.ws.cell(row, h["tags"]).value): flags.append("missing_tags")
        ko = norm(self.ws.cell(row, h["example_ko"]).value)
        if ko and len(ko) < 8: flags.append("short_translation")
        return ", ".join(flags)

    def is_complete_row(self, row): return not self.qa_flag_for_row(row)

    def update_current(self, example_ko, memo, level, tags):
        row = self.data_rows[self.current_pos]
        h = self.headers
        self.ws.cell(row, h["example_ko"]).value = example_ko.strip()
        self.ws.cell(row, h["memo"]).value = memo.strip()
        self.ws.cell(row, h["level"]).value = level.strip()
        self.ws.cell(row, h["tags"]).value = tags.strip()
        self.ws.cell(row, h["qa_flag"]).value = self.qa_flag_for_row(row)

    def qa_summary_for_rows(self, rows):
        summary = {"total":0,"complete":0,"missing_example_ko":0,"missing_memo":0,"missing_level":0,"missing_tags":0,"short_translation":0}
        for row in rows:
            summary["total"] += 1
            flag = self.qa_flag_for_row(row)
            self.ws.cell(row, self.headers["qa_flag"]).value = flag
            if not flag: summary["complete"] += 1
            for k in list(summary.keys()):
                if k not in ("total","complete") and k in flag: summary[k] += 1
        return summary

    def backup(self):
        backup_dir = self.path.parent / "backups"
        backup_dir.mkdir(exist_ok=True)
        backup_path = backup_dir / f"{self.path.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{self.path.suffix}"
        shutil.copy2(self.path, backup_path)
        return backup_path

    def append_log(self, note):
        ws = self.wb["Translation_Log"] if "Translation_Log" in self.wb.sheetnames else self.wb.create_sheet("Translation_Log")
        if ws.max_row == 1 and not ws.cell(1,1).value:
            ws.append(["timestamp","tool","note"])
        ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "dictionary_editor_qt_v2.5", note])

    def save(self):
        self.style()
        self.append_log("Saved from Editor v2.5 Production Mode")
        self.wb.save(self.path)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("DailyVocab Master Dictionary Editor v2.5 Production Mode")
        self.resize(1520, 880)
        self.model = DictionaryWorkbook()
        self.engine = KnowledgeEngine()
        self.visible_rows: List[int] = []
        self.dirty = False
        self.build_actions()
        self.build_ui()
        self.setStyleSheet(DARK_STYLE)
        self.statusBar().showMessage("Excel 파일을 열어 주세요.")

    def build_actions(self):
        save = QAction("Save", self); save.setShortcut(QKeySequence.Save); save.triggered.connect(self.save_file); self.addAction(save)

        save_next = QAction("Save Next", self)
        save_next.setShortcut("Ctrl+Return")
        save_next.triggered.connect(self.save_and_next)
        self.addAction(save_next)

        assist_all = QAction("Assist All", self)
        assist_all.setShortcut("Ctrl+Space")
        assist_all.triggered.connect(self.assist_all)
        self.addAction(assist_all)

        qa = QAction("QA Check", self)
        qa.setShortcut("Ctrl+Q")
        qa.triggered.connect(self.run_qa)
        self.addAction(qa)

        golden = QAction("Add Golden", self)
        golden.setShortcut("Ctrl+B")
        golden.triggered.connect(self.add_golden_current)
        self.addAction(golden)

    def section(self, text):
        label = QLabel(text); label.setObjectName("SectionTitle"); return label
    def editor(self, readonly=False):
        e = QPlainTextEdit(); e.setReadOnly(readonly); e.setLineWrapMode(QPlainTextEdit.WidgetWidth); return e
    def mark_dirty(self):
        if self.model.path:
            self.dirty = True; self.save_state_label.setText("● Unsaved")

    def build_ui(self):
        central = QWidget(); root = QVBoxLayout(central); root.setContentsMargins(14,14,14,10)
        toolbar = QHBoxLayout()
        self.open_btn = QPushButton("Excel 열기"); self.open_btn.clicked.connect(self.open_file)
        self.save_btn = QPushButton("저장"); self.save_btn.clicked.connect(self.save_file)
        self.qa_btn = QPushButton("QA 검사"); self.qa_btn.setObjectName("Secondary"); self.qa_btn.clicked.connect(self.run_qa)
        toolbar.addWidget(self.open_btn); toolbar.addWidget(self.save_btn); toolbar.addWidget(self.qa_btn)
        toolbar.addSpacing(12); toolbar.addWidget(QLabel("행 이동"))
        self.goto_input = QLineEdit(); self.goto_input.setFixedWidth(90)
        self.goto_btn = QPushButton("Go"); self.goto_btn.clicked.connect(self.goto_row)
        toolbar.addWidget(self.goto_input); toolbar.addWidget(self.goto_btn)
        self.save_state_label = QLabel("● No file"); toolbar.addWidget(self.save_state_label)
        root.addLayout(toolbar)

        top = QFrame(); top.setObjectName("Card"); top_layout = QGridLayout(top)
        self.row_label = QLabel("Row -")
        self.word_label = QLabel("No word"); self.word_label.setObjectName("TitleWord")
        self.meaning_label = QLabel(""); self.meaning_label.setObjectName("Meaning")
        self.progress = QProgressBar(); self.progress.setRange(0,100)
        self.progress_label = QLabel("0 / 0")
        top_layout.addWidget(self.row_label,0,0); top_layout.addWidget(self.progress_label,0,1,alignment=Qt.AlignRight)
        top_layout.addWidget(self.word_label,1,0); top_layout.addWidget(self.meaning_label,2,0); top_layout.addWidget(self.progress,1,1,2,1)
        root.addWidget(top)

        splitter = QSplitter(Qt.Horizontal); root.addWidget(splitter,1)

        left = QFrame(); left.setObjectName("Card"); left_l = QVBoxLayout(left)
        left_l.addWidget(self.section("Batch Range"))
        br = QHBoxLayout(); self.batch_start = QLineEdit("1002"); self.batch_end = QLineEdit("1101")
        self.apply_batch_btn = QPushButton("적용"); self.apply_batch_btn.clicked.connect(self.apply_batch_filter)
        br.addWidget(self.batch_start); br.addWidget(QLabel("~")); br.addWidget(self.batch_end); br.addWidget(self.apply_batch_btn); left_l.addLayout(br)
        self.incomplete_only = QCheckBox("미완료만 보기"); self.incomplete_only.stateChanged.connect(self.apply_batch_filter); left_l.addWidget(self.incomplete_only)
        self.auto_assist = QCheckBox("Auto Assist")
        self.auto_assist.setChecked(True)
        left_l.addWidget(self.auto_assist)
        self.word_list = QListWidget(); self.word_list.itemClicked.connect(self.open_word_list_item); left_l.addWidget(self.word_list,1)
        self.dashboard = QPlainTextEdit(); self.dashboard.setReadOnly(True); self.dashboard.setMaximumHeight(140); left_l.addWidget(self.dashboard)

        center = QFrame(); center.setObjectName("Card"); center_l = QVBoxLayout(center)
        self.example_text = self.editor(readonly=True); self.example_ko_text = self.editor(); self.memo_text = self.editor()
        center_l.addWidget(self.section("Example")); center_l.addWidget(self.example_text,2)
        center_l.addWidget(self.section("Example_KO")); center_l.addWidget(self.example_ko_text,2)
        center_l.addWidget(self.section("Memo")); center_l.addWidget(self.memo_text,2)
        nav = QHBoxLayout()
        self.prev_btn = QPushButton("← 이전"); self.prev_btn.setObjectName("Secondary"); self.prev_btn.clicked.connect(self.prev_row)
        self.save_next_btn = QPushButton("저장 후 다음 →"); self.save_next_btn.clicked.connect(self.save_and_next)
        self.next_btn = QPushButton("다음 →"); self.next_btn.setObjectName("Secondary"); self.next_btn.clicked.connect(self.next_row)
        nav.addWidget(self.prev_btn); nav.addWidget(self.save_next_btn); nav.addWidget(self.next_btn); center_l.addLayout(nav)

        right = QFrame(); right.setObjectName("Card"); right_l = QVBoxLayout(right)
        right_l.addWidget(self.section("Level")); self.level_combo = QComboBox(); self.level_combo.addItems(LEVEL_OPTIONS); self.level_combo.currentTextChanged.connect(self.mark_dirty); right_l.addWidget(self.level_combo)
        right_l.addWidget(self.section("Tags")); self.tags_input = QLineEdit(); self.tags_input.textChanged.connect(self.mark_dirty); right_l.addWidget(self.tags_input)
        right_l.addWidget(self.section("Knowledge Assist"))
        self.assist_ko_btn = QPushButton("Example_KO 초안"); self.assist_ko_btn.setObjectName("Assist"); self.assist_ko_btn.clicked.connect(self.assist_example_ko)
        self.assist_memo_btn = QPushButton("Memo 생성"); self.assist_memo_btn.setObjectName("Assist"); self.assist_memo_btn.clicked.connect(self.assist_memo)
        self.assist_tags_btn = QPushButton("Tags 추천"); self.assist_tags_btn.setObjectName("Assist"); self.assist_tags_btn.clicked.connect(self.assist_tags)
        self.assist_level_btn = QPushButton("Level 추천"); self.assist_level_btn.setObjectName("Assist"); self.assist_level_btn.clicked.connect(self.assist_level)
        self.assist_all_btn = QPushButton("Assist All"); self.assist_all_btn.setObjectName("Assist"); self.assist_all_btn.clicked.connect(self.assist_all)
        for b in [self.assist_ko_btn,self.assist_memo_btn,self.assist_tags_btn,self.assist_level_btn,self.assist_all_btn]: right_l.addWidget(b)
        self.golden_btn = QPushButton("⭐ Golden 저장")
        self.golden_btn.setObjectName("Assist")
        self.golden_btn.clicked.connect(self.add_golden_current)
        right_l.addWidget(self.golden_btn)
        right_l.addWidget(self.section("Confidence")); self.confidence_label = QLabel("No suggestion yet"); self.confidence_label.setWordWrap(True); right_l.addWidget(self.confidence_label)
        right_l.addWidget(self.section("Suggestion Reasons")); self.reasons_text = QPlainTextEdit(); self.reasons_text.setReadOnly(True); self.reasons_text.setMaximumHeight(95); right_l.addWidget(self.reasons_text)
        right_l.addWidget(self.section("QA Flag")); self.qa_flag_label = QLabel("OK"); self.qa_flag_label.setWordWrap(True); right_l.addWidget(self.qa_flag_label)
        right_l.addWidget(self.section("QA Summary")); self.qa_summary_text = QPlainTextEdit(); self.qa_summary_text.setReadOnly(True); right_l.addWidget(self.qa_summary_text,1)

        splitter.addWidget(left); splitter.addWidget(center); splitter.addWidget(right)
        splitter.setStretchFactor(0,1); splitter.setStretchFactor(1,3); splitter.setStretchFactor(2,2)
        self.setCentralWidget(central); self.setStatusBar(QStatusBar())
        self.example_ko_text.textChanged.connect(self.mark_dirty); self.memo_text.textChanged.connect(self.mark_dirty)

    def open_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "Excel 파일 열기", "", "Excel Files (*.xlsx)")
        if not path: return
        self.model.load(Path(path)); self.visible_rows = list(self.model.data_rows)
        self.load_current_to_ui(); self.apply_batch_filter()
        self.save_state_label.setText("● Saved"); self.statusBar().showMessage(f"Loaded: {Path(path).name}")

    def apply_batch_filter(self):
        if not self.model.data_rows: return
        try: start, end = int(self.batch_start.text()), int(self.batch_end.text())
        except ValueError:
            QMessageBox.warning(self,"Batch Range","시작/끝 행 번호를 숫자로 입력하세요."); return
        rows = [r for r in self.model.data_rows if start <= r <= end]
        if self.incomplete_only.isChecked(): rows = [r for r in rows if not self.model.is_complete_row(r)]
        self.visible_rows = rows; self.refresh_word_list()
        if rows:
            pos = self.model.pos_for_row(rows[0])
            if pos is not None: self.model.goto_pos(pos); self.load_current_to_ui()
        self.statusBar().showMessage(f"Batch 적용 완료: {len(rows)}개 표시")

    def refresh_word_list(self):
        self.word_list.clear(); complete = 0
        for row_no in self.visible_rows:
            pos = self.model.pos_for_row(row_no)
            if pos is None: continue
            old = self.model.current_pos; self.model.current_pos = pos; row = self.model.get_current(); self.model.current_pos = old
            ok = self.model.is_complete_row(row_no); complete += 1 if ok else 0
            item = QListWidgetItem(f"{'✓' if ok else '•'} {row['row']}  {row['word']}  |  {row['meaning']}")
            item.setData(Qt.UserRole, row_no); self.word_list.addItem(item)
        total = len(self.visible_rows); rate = int(complete/total*100) if total else 0
        self.dashboard.setPlainText(f"Batch Dashboard\n---------------\nVisible: {total}\nComplete: {complete}\nRemaining: {total-complete}\nRate: {rate}%")

    def maybe_auto_assist(self):
        if not hasattr(self, "auto_assist") or not self.auto_assist.isChecked():
            return
        if not self.model.data_rows:
            return
        # Only fill empty fields automatically. Existing user content is preserved.
        row = self.model.get_current()
        s = self.suggestion()
        changed = False
        if not self.example_ko_text.toPlainText().strip():
            self.example_ko_text.setPlainText(s.example_ko)
            changed = True
        if not self.memo_text.toPlainText().strip():
            self.memo_text.setPlainText(s.memo)
            changed = True
        if not self.tags_input.text().strip():
            self.tags_input.setText(s.tags)
            changed = True
        if not self.level_combo.currentText().strip():
            self.level_combo.setCurrentText(s.level)
            changed = True
        if changed:
            self.mark_dirty()

    def add_golden_current(self):
        if not self.model.data_rows:
            return
        row = self.model.get_current()
        ko = self.example_ko_text.toPlainText().strip()
        memo = self.memo_text.toPlainText().strip()
        try:
            if ko:
                self.engine.add_golden("translations", row["example"], ko, f"{row['word']} / {row['meaning']}")
            if memo:
                self.engine.add_golden("memos", row["word"], memo, row["example"])
            self.statusBar().showMessage("⭐ Golden Collection에 저장했습니다.")
        except Exception as e:
            QMessageBox.warning(self, "Golden 저장", f"저장 중 오류가 발생했습니다: {e}")

    def load_current_to_ui(self):
        row = self.model.get_current()
        self.row_label.setText(f"Excel Row {row['row']}"); self.progress_label.setText(f"{self.model.current_pos+1} / {self.model.count()}")
        self.progress.setValue(int((self.model.current_pos+1)/self.model.count()*100))
        self.word_label.setText(row["word"]); self.meaning_label.setText(row["meaning"])
        self.example_text.setPlainText(row["example"])
        self.example_ko_text.blockSignals(True); self.memo_text.blockSignals(True)
        self.example_ko_text.setPlainText(row["example_ko"]); self.memo_text.setPlainText(row["memo"])
        self.example_ko_text.blockSignals(False); self.memo_text.blockSignals(False)
        self.level_combo.blockSignals(True); self.level_combo.setCurrentText(row["level"] if row["level"] in LEVEL_OPTIONS else ""); self.level_combo.blockSignals(False)
        self.tags_input.blockSignals(True); self.tags_input.setText(row["tags"]); self.tags_input.blockSignals(False)
        flag = row["qa_flag"] or self.model.qa_flag_for_row(row["row"]) or "OK"
        self.qa_flag_label.setText(flag); self.qa_flag_label.setStyleSheet("color:#34D399;" if flag == "OK" else "color:#F87171;")
        self.goto_input.setText(str(row["row"]))
        self.confidence_label.setText("No suggestion yet"); self.confidence_label.setStyleSheet(""); self.reasons_text.setPlainText("")
        self.dirty = False; self.save_state_label.setText("● Saved")
        self.maybe_auto_assist()

    def suggestion(self):
        row = self.model.get_current()
        s = self.engine.suggest_all(row["word"], row["meaning"], row["example"])
        if s.confidence >= 85: color, label = "#34D399", "🟢 High"
        elif s.confidence >= 65: color, label = "#FBBF24", "🟡 Medium"
        else: color, label = "#F87171", "🔴 Review"
        self.confidence_label.setText(f"{label} Confidence: {s.confidence}%")
        self.confidence_label.setStyleSheet(f"color:{color}; font-weight:700;")
        self.reasons_text.setPlainText("\n".join(s.reasons) if s.reasons else "No specific reason.")
        return s

    def assist_example_ko(self):
        if self.example_ko_text.toPlainText().strip():
            r = QMessageBox.question(self,"Example_KO 초안","이미 내용이 있습니다. 초안으로 덮어쓸까요?", QMessageBox.Yes|QMessageBox.No)
            if r != QMessageBox.Yes: return
        self.example_ko_text.setPlainText(self.suggestion().example_ko)

    def assist_memo(self): self.memo_text.setPlainText(self.suggestion().memo)
    def assist_tags(self): self.tags_input.setText(self.suggestion().tags)
    def assist_level(self): self.level_combo.setCurrentText(self.suggestion().level)
    def assist_all(self):
        s = self.suggestion()
        if not self.example_ko_text.toPlainText().strip(): self.example_ko_text.setPlainText(s.example_ko)
        self.memo_text.setPlainText(s.memo); self.tags_input.setText(s.tags); self.level_combo.setCurrentText(s.level)

    def save_current_from_ui(self):
        row = self.model.get_current()
        ko = self.example_ko_text.toPlainText().strip()
        self.model.update_current(ko, self.memo_text.toPlainText(), self.level_combo.currentText(), self.tags_input.text())
        if ko:
            try:
                self.engine.learn_translation(row["example"], ko, confidence=88)
                for phrase, info in self.engine.find_phrase_matches(row["example"]):
                    self.engine.learn_phrase(phrase, info.get("ko",""), info.get("note",""), info.get("confidence",95))
            except Exception:
                pass
        flag = self.model.qa_flag_for_row(row["row"]) or "OK"
        self.qa_flag_label.setText(flag); self.qa_flag_label.setStyleSheet("color:#34D399;" if flag == "OK" else "color:#F87171;")
        self.dirty = False; self.save_state_label.setText("● Saved"); self.refresh_word_list()

    def open_word_list_item(self, item):
        if self.dirty: self.save_current_from_ui()
        self.model.goto_row_number(item.data(Qt.UserRole)); self.load_current_to_ui()
    def prev_row(self):
        if self.dirty: self.save_current_from_ui()
        self.model.goto_pos(self.model.current_pos-1); self.load_current_to_ui()
    def next_row(self):
        if self.dirty: self.save_current_from_ui()
        self.model.goto_pos(self.model.current_pos+1); self.load_current_to_ui()
    def save_and_next(self):
        self.save_current_from_ui()
        current = self.model.get_current()["row"]
        if self.visible_rows and current in self.visible_rows:
            idx = self.visible_rows.index(current)
            self.model.goto_row_number(self.visible_rows[min(idx+1, len(self.visible_rows)-1)])
        else:
            self.model.goto_pos(self.model.current_pos+1)
        self.load_current_to_ui()
    def goto_row(self):
        if self.dirty: self.save_current_from_ui()
        try: row_no = int(self.goto_input.text())
        except ValueError:
            QMessageBox.warning(self,"행 이동","행 번호를 숫자로 입력하세요."); return
        self.model.goto_row_number(row_no); self.load_current_to_ui()
    def run_qa(self):
        self.save_current_from_ui()
        rows = self.visible_rows if self.visible_rows else self.model.data_rows
        s = self.model.qa_summary_for_rows(rows); rate = int(s["complete"]/s["total"]*100) if s["total"] else 0
        self.qa_summary_text.setPlainText(
            f"QA Summary\n----------\nScope rows: {s['total']}\nComplete: {s['complete']}\nMissing example_ko: {s['missing_example_ko']}\nMissing memo: {s['missing_memo']}\nMissing level: {s['missing_level']}\nMissing tags: {s['missing_tags']}\nShort translation: {s['short_translation']}\nComplete rate: {rate}%"
        )
        self.progress.setValue(rate); self.refresh_word_list(); self.statusBar().showMessage("QA scan complete")
    def save_file(self):
        if not self.model.path:
            QMessageBox.warning(self,"저장","먼저 Excel 파일을 열어 주세요."); return
        self.save_current_from_ui(); backup = self.model.backup(); self.model.save()
        self.save_state_label.setText("● Saved"); self.statusBar().showMessage(f"Saved. Backup: {backup.name}")

def main():
    app = QApplication(sys.argv)
    win = MainWindow(); win.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
