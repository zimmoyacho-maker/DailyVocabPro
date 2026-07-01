from openpyxl import load_workbook
from app.core.paths import words_xlsx_path
from app.database.db import connect

REQUIRED = ["word", "meaning", "example"]
OPTIONAL = ["example_ko", "memo", "level", "tags"]
ALL = REQUIRED + OPTIONAL

def import_words_from_excel() -> int:
    path = words_xlsx_path()
    if not path.exists():
        raise FileNotFoundError(f"words.xlsx not found: {path}")

    wb = load_workbook(path)
    ws = wb["words"] if "words" in wb.sheetnames else wb.active

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    header_map = {h: i for i, h in enumerate(headers)}

    missing = [h for h in REQUIRED if h not in header_map]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {}
        for col in ALL:
            idx = header_map.get(col)
            value = row[idx] if idx is not None and idx < len(row) else ""
            item[col] = str(value).strip() if value is not None else ""
        if item["word"]:
            rows.append(item)

    conn = connect()
    cur = conn.cursor()

    for item in rows:
        cur.execute("""
        INSERT INTO words(word, meaning, example, example_ko, memo, level, tags)
        VALUES(?,?,?,?,?,?,?)
        ON CONFLICT(word) DO UPDATE SET
            meaning=excluded.meaning,
            example=excluded.example,
            example_ko=excluded.example_ko,
            memo=excluded.memo,
            level=excluded.level,
            tags=excluded.tags
        """, (
            item["word"],
            item["meaning"],
            item["example"],
            item["example_ko"],
            item["memo"],
            item["level"],
            item["tags"],
        ))

    conn.commit()
    conn.close()
    return len(rows)
