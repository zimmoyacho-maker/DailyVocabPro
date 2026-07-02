#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DailyVocabPro Master Dictionary Builder

Purpose
-------
Fill/assist columns in words.xlsx:
- example_ko
- memo
- level
- tags
- optional dictionary columns

External sources
----------------
1) Free Dictionary API
   https://api.dictionaryapi.dev/api/v2/entries/en/{word}

2) LibreTranslate-compatible API
   Default: http://localhost:5000/translate
   Docs: https://libretranslate.com/docs/

Recommended flow
----------------
1. Run a small test batch first.
2. Review output.
3. Run 100~500 rows at a time.
4. Manually review suspicious rows using QA sheet.

Examples
--------
# Local LibreTranslate server
python master_dictionary_builder.py --input words.xlsx --output words_out.xlsx --start-row 1002 --end-row 1101 --translate-url http://localhost:5000/translate

# LibreTranslate cloud/server that requires an API key
python master_dictionary_builder.py --input words.xlsx --output words_out.xlsx --start-row 1002 --end-row 1101 --translate-url https://libretranslate.com/translate --api-key YOUR_KEY

# Dictionary/memo/tags only, no translation API
python master_dictionary_builder.py --input words.xlsx --output words_out.xlsx --start-row 1002 --end-row 1101 --no-translate
"""

from __future__ import annotations

import argparse
import json
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


DEFAULT_TRANSLATE_URL = "http://localhost:5000/translate"
FREE_DICTIONARY_URL = "https://api.dictionaryapi.dev/api/v2/entries/en/{word}"


REQUIRED_COLUMNS = ["word", "meaning", "example"]
BASE_COLUMNS = ["word", "meaning", "example", "example_ko", "memo", "level", "tags"]
EXTRA_COLUMNS = [
    "part_of_speech",
    "dictionary_def",
    "pronunciation",
    "source_url",
    "qa_flag",
]


COMMON_A1 = {
    "i", "you", "he", "she", "we", "they", "it", "be", "is", "are", "am", "was", "were",
    "have", "has", "do", "go", "come", "get", "make", "take", "see", "look", "want",
    "need", "like", "good", "bad", "big", "small", "new", "old", "hot", "cold", "water",
    "food", "house", "home", "school", "book", "car", "bus", "train", "dog", "cat",
    "man", "woman", "child", "day", "night", "morning", "time", "year", "today",
}
COMMON_A2 = {
    "traffic", "wall", "disease", "calendar", "temperature", "exercise", "damage",
    "floor", "roof", "coast", "route", "rain", "purpose", "change", "escape", "ready",
    "population", "poor", "arrest", "join", "score", "plan", "rabbit", "area", "record",
    "university", "shade", "govern", "tower", "realize", "correct", "start", "movie",
    "entertain", "cage", "ocean", "food", "inside", "order", "guest", "guide",
    "control", "save", "red", "gate", "plant", "draw", "clay", "green", "flavor",
    "wave", "break", "crop", "fall", "output", "authentic", "east", "hall", "drain",
    "dog", "school", "dive", "ceiling", "delay", "death", "cycle", "blood", "honor",
    "scatter", "clue", "corner", "core", "conflict", "exceed", "policy", "explore",
    "prison", "bitter", "weapon", "cinema", "cap", "monster", "weak", "cheap",
    "unit", "huge", "sea", "finish", "captain", "sand", "waste", "pork", "brown",
    "broad", "branch", "blue", "lap", "accept", "raise", "story", "recommend",
    "army", "liberty", "award", "each", "angle", "amount", "moon", "sky", "bear",
    "century", "bite", "weave", "environment", "face", "generous", "progress",
}


TAG_KEYWORDS = {
    "Business": [
        "company", "market", "client", "server", "contract", "revenue", "trading",
        "business", "customer", "sales", "office", "policy", "committee", "bank",
        "price", "cost", "fund", "proposal", "manager",
    ],
    "Academic": [
        "theory", "philosophy", "research", "scientist", "mathematical", "equation",
        "vector", "dimension", "unit", "definition", "language", "university",
        "college", "student", "teacher", "school", "pupil", "field",
    ],
    "Science": [
        "plant", "temperature", "biology", "geology", "disease", "drug", "velocity",
        "microscope", "fusion", "laboratory", "chemical", "species", "cell",
    ],
    "Medical": [
        "doctor", "medicine", "medication", "disease", "health", "hospital", "drug",
        "cancer", "patient", "pregnancy", "pain", "aspirin",
    ],
    "Legal": [
        "court", "law", "police", "crime", "murder", "arrest", "custody", "defendant",
        "judge", "sentence", "legal", "complainant", "damages",
    ],
    "Travel": [
        "airport", "hotel", "restaurant", "ship", "train", "bus", "station", "route",
        "travel", "passport", "embassy", "tour", "museum", "harbor", "island",
    ],
    "Technology": [
        "software", "computer", "server", "protocol", "network", "program", "http",
        "windows", "device", "data", "scan", "file",
    ],
    "Daily": [
        "house", "room", "family", "food", "coffee", "clothes", "baby", "dog", "cat",
        "friend", "home", "desk", "kitchen", "bathroom", "toilet",
    ],
}


COLLOCATION_PATTERNS = [
    (r"\bbe obsessed with\b", "be obsessed with = ~에 집착하다 / 푹 빠져 있다"),
    (r"\bbe based on\b", "be based on = ~에 근거하다"),
    (r"\bbe famous for\b", "be famous for = ~로 유명하다"),
    (r"\bbe famous as\b", "be famous as = ~로서 유명하다"),
    (r"\btake (a|the)?\s?register\b", "take the register = 출석을 부르다"),
    (r"\bmake a decision\b", "make a decision = 결정하다"),
    (r"\btake responsibility\b", "take responsibility = 책임지다"),
    (r"\bin progress\b", "in progress = 진행 중인"),
    (r"\bdepend(?:s|ed|ing)? upon\b", "depend upon = ~에 달려 있다"),
    (r"\bconsist(?:s|ed|ing)? of\b", "consist of = ~로 이루어져 있다"),
    (r"\bprohibit(?:s|ed|ing)?\b", "prohibit + 목적어/from -ing = ~을 금지하다"),
    (r"\bbe charged with\b", "be charged with = ~혐의로 기소되다"),
]


@dataclass
class DictionaryInfo:
    part_of_speech: str = ""
    definition: str = ""
    pronunciation: str = ""
    source_url: str = ""


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").replace("\u2003", " ").strip()


def load_cache(cache_path: Path) -> Dict[str, Any]:
    if cache_path.exists():
        try:
            return json.loads(cache_path.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_cache(cache_path: Path, cache: Dict[str, Any]) -> None:
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def request_json(url: str, *, method: str = "GET", payload: Optional[dict] = None, timeout: int = 20) -> Any:
    if method.upper() == "POST":
        res = requests.post(url, json=payload, timeout=timeout)
    else:
        res = requests.get(url, timeout=timeout)
    res.raise_for_status()
    return res.json()


def fetch_dictionary(word: str, cache: Dict[str, Any], sleep_sec: float = 0.1) -> DictionaryInfo:
    key = f"dict:{word.lower()}"
    if key in cache:
        return parse_dictionary(word, cache[key])

    url = FREE_DICTIONARY_URL.format(word=quote(word.lower()))
    try:
        data = request_json(url)
        cache[key] = data
        time.sleep(sleep_sec)
        return parse_dictionary(word, data)
    except Exception as exc:
        cache[key] = {"error": str(exc)}
        return DictionaryInfo(source_url=url)


def parse_dictionary(word: str, data: Any) -> DictionaryInfo:
    url = FREE_DICTIONARY_URL.format(word=quote(word.lower()))
    if isinstance(data, dict) and "error" in data:
        return DictionaryInfo(source_url=url)

    if not isinstance(data, list) or not data:
        return DictionaryInfo(source_url=url)

    entry = data[0]
    pronunciation = ""
    for p in entry.get("phonetics", []) or []:
        if p.get("text"):
            pronunciation = p.get("text", "")
            break

    meanings = entry.get("meanings", []) or []
    if not meanings:
        return DictionaryInfo(pronunciation=pronunciation, source_url=url)

    first = meanings[0]
    pos = first.get("partOfSpeech", "") or ""
    definitions = first.get("definitions", []) or []
    definition = ""
    if definitions:
        definition = definitions[0].get("definition", "") or ""

    return DictionaryInfo(
        part_of_speech=pos,
        definition=definition,
        pronunciation=pronunciation,
        source_url=url,
    )


def translate_text(
    text: str,
    translate_url: str,
    api_key: str = "",
    cache: Optional[Dict[str, Any]] = None,
    sleep_sec: float = 0.2,
) -> str:
    if not text:
        return ""

    key = f"translate:en-ko:{text}"
    if cache is not None and key in cache:
        return cache[key]

    payload = {
        "q": text,
        "source": "en",
        "target": "ko",
        "format": "text",
    }
    if api_key:
        payload["api_key"] = api_key

    try:
        data = request_json(translate_url, method="POST", payload=payload)
        translated = norm(data.get("translatedText", ""))
        if cache is not None:
            cache[key] = translated
        time.sleep(sleep_sec)
        return translated
    except Exception as exc:
        if cache is not None:
            cache[key] = {"error": str(exc)}
        return ""


def polish_korean_translation(word: str, meaning: str, example: str, raw_ko: str) -> str:
    """Light polishing for machine translation output.

    This does not pretend to be a human translation. It removes common awkward endings
    and returns a clean draft for manual review.
    """
    ko = norm(raw_ko)
    if not ko:
        return ""

    replacements = {
        "이다.": "입니다.",
        "한다.": "합니다.",
        "된다.": "됩니다.",
        "했다.": "했습니다.",
        "되었다.": "되었습니다.",
        "있다.": "있습니다.",
        "없다.": "없습니다.",
        "있었다.": "있었습니다.",
        "없었다.": "없었습니다.",
    }
    for src, dst in replacements.items():
        if ko.endswith(src):
            ko = ko[: -len(src)] + dst
            break

    # Prevent suspicious ultra-short outputs
    if len(ko) < 6:
        return f"{ko}  [검수 필요]"
    return ko


def infer_level(word: str, meaning: str, example: str, dict_info: DictionaryInfo) -> str:
    w = word.lower()
    e = example.lower()

    if w in COMMON_A1:
        return "A1"
    if w in COMMON_A2:
        return "A2"

    if dict_info.part_of_speech in {"conjunction", "preposition", "determiner", "pronoun"} and len(w) <= 6:
        return "A2"

    # technical / abstract clues
    if any(k in e for k in ["equation", "theory", "vector", "protocol", "political", "government", "dependent", "philosophy"]):
        return "B2"

    if len(w) >= 12 or len(example) >= 150:
        return "C1"
    if len(w) >= 9 or len(example) >= 90:
        return "B2"
    return "B1"


def infer_tags(word: str, meaning: str, example: str, dict_info: DictionaryInfo) -> str:
    text = " ".join([word, meaning, example, dict_info.definition]).lower()
    tags: List[str] = []
    for tag, keywords in TAG_KEYWORDS.items():
        if any(k in text for k in keywords):
            tags.append(tag)

    if not tags:
        tags = ["Daily"]

    # Limit to 2 tags for readability
    return ", ".join(tags[:2])


def detect_collocation(example: str) -> str:
    e = example.lower()
    matches = []
    for pattern, label in COLLOCATION_PATTERNS:
        if re.search(pattern, e):
            matches.append(label)
    return " / ".join(matches)


def build_memo(word: str, meaning: str, example: str, dict_info: DictionaryInfo) -> str:
    collocation = detect_collocation(example)
    pos = dict_info.part_of_speech
    definition = dict_info.definition

    pieces: List[str] = []

    if collocation:
        pieces.append(collocation)

    # Add high-value pattern notes
    lower_example = example.lower()
    if "with" in lower_example and word.lower() in lower_example:
        if not collocation:
            pieces.append("전치사 with와 함께 쓰이는지 확인해 보세요.")
    if "of" in lower_example and word.lower() in lower_example:
        pieces.append("of와 함께 쓰일 때의 의미 확장을 확인하세요.")

    if pos:
        pieces.append(f"품사: {pos}")

    if definition:
        short_def = definition
        if len(short_def) > 110:
            short_def = short_def[:107].rstrip() + "..."
        pieces.append(f"영영 뜻: {short_def}")

    if not pieces:
        pieces.append(f"'{word}'는 '{meaning}'의 의미로 예문 속 문맥과 함께 기억하세요.")

    return " ".join(pieces)


def qa_flag(example_ko: str, memo: str, level: str, tags: str, used_translation_api: bool) -> str:
    flags = []
    if not example_ko:
        flags.append("missing_example_ko")
    if example_ko and ("검수 필요" in example_ko or len(example_ko) < 8):
        flags.append("review_translation")
    if not memo:
        flags.append("missing_memo")
    if not level:
        flags.append("missing_level")
    if not tags:
        flags.append("missing_tags")
    if not used_translation_api:
        flags.append("no_translation_api")
    return ", ".join(flags)


def get_header_map(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = norm(ws.cell(row=1, column=col).value)
        if value:
            headers[value] = col
    return headers


def ensure_columns(ws, columns: List[str]) -> Dict[str, int]:
    headers = get_header_map(ws)
    next_col = ws.max_column + 1

    for col_name in columns:
        if col_name not in headers:
            ws.cell(row=1, column=next_col).value = col_name
            headers[col_name] = next_col
            next_col += 1

    return headers


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "word": 18,
        "meaning": 24,
        "example": 50,
        "example_ko": 50,
        "memo": 56,
        "level": 10,
        "tags": 18,
        "part_of_speech": 16,
        "dictionary_def": 52,
        "pronunciation": 18,
        "source_url": 46,
        "qa_flag": 28,
    }
    headers = get_header_map(ws)
    for name, col in headers.items():
        ws.column_dimensions[get_column_letter(col)].width = widths.get(name, 18)

    ws.freeze_panes = "A2"
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def append_log(wb, batch_name: str, start_row: int, end_row: int, notes: str) -> None:
    ws = wb["Translation_Log"] if "Translation_Log" in wb.sheetnames else wb.create_sheet("Translation_Log")
    if ws.max_row == 1 and not ws.cell(1, 1).value:
        ws.append(["timestamp", "batch", "range", "notes"])
    ws.append([
        time.strftime("%Y-%m-%d %H:%M:%S"),
        batch_name,
        f"Sheet1!{start_row}:{end_row}",
        notes,
    ])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.font = Font(bold=True, color="FFFFFF")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 80


def build_dictionary(args: argparse.Namespace) -> None:
    input_path = Path(args.input)
    output_path = Path(args.output)
    cache_path = Path(args.cache)

    cache = load_cache(cache_path)

    wb = load_workbook(input_path)
    ws = wb[args.sheet] if args.sheet else wb.active

    headers = ensure_columns(ws, BASE_COLUMNS + EXTRA_COLUMNS)
    for required in REQUIRED_COLUMNS:
        if required not in headers:
            raise ValueError(f"Required column missing: {required}")

    processed = 0
    translated_count = 0
    dictionary_count = 0

    for row_idx in range(args.start_row, args.end_row + 1):
        word = norm(ws.cell(row=row_idx, column=headers["word"]).value)
        meaning = norm(ws.cell(row=row_idx, column=headers["meaning"]).value)
        example = norm(ws.cell(row=row_idx, column=headers["example"]).value)

        if not word:
            continue

        existing_ko = norm(ws.cell(row=row_idx, column=headers["example_ko"]).value)
        existing_memo = norm(ws.cell(row=row_idx, column=headers["memo"]).value)
        existing_level = norm(ws.cell(row=row_idx, column=headers["level"]).value)
        existing_tags = norm(ws.cell(row=row_idx, column=headers["tags"]).value)

        dict_info = fetch_dictionary(word, cache, sleep_sec=args.sleep)
        if dict_info.definition or dict_info.part_of_speech:
            dictionary_count += 1

        used_translation_api = False
        example_ko = existing_ko

        if (args.overwrite or not example_ko) and not args.no_translate:
            raw_ko = translate_text(
                example,
                args.translate_url,
                api_key=args.api_key,
                cache=cache,
                sleep_sec=args.sleep,
            )
            if raw_ko:
                used_translation_api = True
                example_ko = polish_korean_translation(word, meaning, example, raw_ko)
                translated_count += 1

        # If translation API unavailable, keep blank and flag for manual work.
        if args.overwrite or not existing_memo:
            memo = build_memo(word, meaning, example, dict_info)
        else:
            memo = existing_memo

        if args.overwrite or not existing_level:
            level = infer_level(word, meaning, example, dict_info)
        else:
            level = existing_level

        if args.overwrite or not existing_tags:
            tags = infer_tags(word, meaning, example, dict_info)
        else:
            tags = existing_tags

        ws.cell(row=row_idx, column=headers["example_ko"]).value = example_ko
        ws.cell(row=row_idx, column=headers["memo"]).value = memo
        ws.cell(row=row_idx, column=headers["level"]).value = level
        ws.cell(row=row_idx, column=headers["tags"]).value = tags
        ws.cell(row=row_idx, column=headers["part_of_speech"]).value = dict_info.part_of_speech
        ws.cell(row=row_idx, column=headers["dictionary_def"]).value = dict_info.definition
        ws.cell(row=row_idx, column=headers["pronunciation"]).value = dict_info.pronunciation
        ws.cell(row=row_idx, column=headers["source_url"]).value = dict_info.source_url
        ws.cell(row=row_idx, column=headers["qa_flag"]).value = qa_flag(
            example_ko, memo, level, tags, used_translation_api or bool(existing_ko)
        )

        processed += 1

        if processed % 25 == 0:
            print(f"Processed {processed} rows...")

    style_sheet(ws)
    append_log(
        wb,
        args.batch_name,
        args.start_row,
        args.end_row,
        (
            f"Processed={processed}, translated={translated_count}, "
            f"dictionary_hits={dictionary_count}, overwrite={args.overwrite}, "
            f"translation_url={args.translate_url if not args.no_translate else 'disabled'}"
        ),
    )

    save_cache(cache_path, cache)
    wb.save(output_path)

    print("\nDone.")
    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")
    print(f"Rows processed: {processed}")
    print(f"Translated: {translated_count}")
    print(f"Dictionary hits: {dictionary_count}")
    print(f"Cache: {cache_path}")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build DailyVocabPro Master Dictionary columns.")
    p.add_argument("--input", required=True, help="Input .xlsx file")
    p.add_argument("--output", required=True, help="Output .xlsx file")
    p.add_argument("--sheet", default="", help="Sheet name. Default: active sheet")
    p.add_argument("--start-row", type=int, required=True, help="Start row number, e.g. 1002")
    p.add_argument("--end-row", type=int, required=True, help="End row number, e.g. 1101")
    p.add_argument("--batch-name", default="Batch", help="Log batch name")
    p.add_argument("--translate-url", default=DEFAULT_TRANSLATE_URL, help="LibreTranslate-compatible /translate endpoint")
    p.add_argument("--api-key", default="", help="LibreTranslate API key if required")
    p.add_argument("--no-translate", action="store_true", help="Disable translation API and only fill memo/level/tags/dictionary info")
    p.add_argument("--overwrite", action="store_true", help="Overwrite existing example_ko/memo/level/tags")
    p.add_argument("--cache", default="master_dictionary_cache.json", help="JSON cache path")
    p.add_argument("--sleep", type=float, default=0.15, help="Delay between API calls")
    return p.parse_args()


if __name__ == "__main__":
    build_dictionary(parse_args())
